#!/usr/bin/env python3
"""
IMPORTADOR DE CATÁLOGO PIEZAUTO - Hito 2 v5.1
Procesa archivos XLSX normalizados del COO y los carga en Supabase

VERSIÓN 5.1: FIX CRÍTICO generador códigos Piezauto
- Cambiado formato PZ-XXXXX (5 dígitos) → PZ-XXXXXX (6 dígitos)
- ORDER BY por longitud primero, luego alfabético (evita bug PZ-100000)
- Batches reducidos (250 SKUs) + delays para evitar pérdida de conexión Supabase

Requisitos:
- openpyxl: lectura de XLSX
- psycopg[binary]: conexión a Postgres/Supabase (versión 3)
- python-dotenv: variables de entorno

Uso:
    python importador_catalogo_v5_1.py --proveedor Cromosol --archivo Cromosol.xlsx
    python importador_catalogo_v5_1.py --modo batch --directorio /path/to/xlsx/
"""

import sys
import argparse
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime
import time
import openpyxl
import psycopg
import os
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('importador.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

COLUMNAS_REQUERIDAS = [
    'codigo_proveedor',
    'descripcion_original',
    'fabricante',
    'familia_sugerida',
    'codigo_oem',
    'precio_lista',
    'precio_neto',
    'aplicaciones',
    'tipo_lados',  # Nueva columna v4
    'lado',
    'observaciones'
]

COLUMNAS_OPCIONALES = [
    'carroceria',
    'puertas',
    'posicion',
    'motor',
    'caja',
    'version',
    'codigo_padre_proveedor'  # Para hijos pre-armados (hijo_de_padre_multiple)
]

# =============================================================================
# MODELOS DE DATOS
# =============================================================================

@dataclass
class SKUNormalizado:
    """Representa un SKU del archivo XLSX normalizado"""
    codigo_proveedor: str
    descripcion_original: str
    fabricante: str
    familia_sugerida: str
    codigo_oem: Optional[str]
    precio_lista: Optional[float]
    precio_neto: Optional[float]
    aplicaciones: str
    tipo_lados: str  # lados_combinados, juego_indivisible, kit, lado_explicito, sin_lado, hijo_de_padre_multiple
    lado: str  # N/A, Der, Izq, Ambos
    carroceria: Optional[str]
    puertas: Optional[str]
    posicion: Optional[str]
    motor: Optional[str]
    caja: Optional[str]
    version: Optional[str]
    observaciones: Optional[str]
    codigo_padre_proveedor: Optional[str]  # Nuevo: para hijos pre-armados (hijo_de_padre_multiple)
    
    # Metadata de procesamiento
    fila_origen: int = 0
    errores_validacion: List[str] = None
    
    def __post_init__(self):
        if self.errores_validacion is None:
            self.errores_validacion = []

@dataclass
class EstadisticasIngesta:
    """Estadísticas del proceso de ingesta"""
    total_filas: int = 0
    validadas: int = 0
    rechazadas: int = 0
    nuevos_skus: int = 0
    skus_actualizados: int = 0
    lados_combinados_padres: int = 0
    lados_combinados_hijos: int = 0
    errores: List[str] = None
    
    def __post_init__(self):
        if self.errores is None:
            self.errores = []

# =============================================================================
# CONEXIÓN A BASE DE DATOS
# =============================================================================

class DatabaseConnection:
    """Manejador de conexión a Supabase/Postgres"""
    
    def __init__(self):
        self.conn = None
        self.cursor = None
        
    def conectar(self):
        """Establece conexión con Supabase"""
        try:
            self.conn = psycopg.connect(
                host=os.getenv('SUPABASE_HOST'),
                dbname=os.getenv('SUPABASE_DB'),
                user=os.getenv('SUPABASE_USER'),
                password=os.getenv('SUPABASE_PASSWORD'),
                port=os.getenv('SUPABASE_PORT', '5432')
            )
            self.cursor = self.conn.cursor()
            logger.info("[OK] Conexion a Supabase establecida")
            return True
        except Exception as e:
            logger.error(f"[ERROR] Error conectando a Supabase: {e}")
            return False
    
    def cerrar(self):
        """Cierra la conexión"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
        logger.info("[OK] Conexion cerrada")
    
    def ejecutar_query(self, query: str, params: tuple = None) -> Optional[List]:
        """Ejecuta una query y devuelve resultados"""
        try:
            self.cursor.execute(query, params)
            if self.cursor.description:  # SELECT
                return self.cursor.fetchall()
            return None
        except Exception as e:
            logger.error(f"[ERROR] Error ejecutando query: {e}")
            raise
    
    def commit(self):
        """Commit de la transacción"""
        self.conn.commit()
    
    def rollback(self):
        """Rollback de la transacción"""
        self.conn.rollback()

# =============================================================================
# ETAPA 1: VALIDACIÓN DE ESTRUCTURA
# =============================================================================

def validar_estructura_xlsx(filepath: Path) -> Tuple[bool, List[str]]:
    """
    Valida que el XLSX tenga la estructura esperada
    
    Returns:
        (es_valido, errores)
    """
    errores = []
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        
        # Verificar hoja "Normalizado"
        if "Normalizado" not in wb.sheetnames:
            errores.append("Falta hoja 'Normalizado'")
            return False, errores
        
        ws = wb["Normalizado"]
        
        # Leer encabezados (primera fila)
        encabezados = [cell.value for cell in ws[1]]
        
        # Verificar columnas requeridas
        for col in COLUMNAS_REQUERIDAS:
            if col not in encabezados:
                errores.append(f"Falta columna requerida: {col}")
        
        if errores:
            return False, errores
        
        # Verificar que haya al menos una fila de datos
        if ws.max_row < 2:
            errores.append("Archivo vacío (sin filas de datos)")
            return False, errores
        
        logger.info(f"[OK] Estructura válida: {ws.max_row - 1} filas de datos")
        return True, []
        
    except Exception as e:
        errores.append(f"Error leyendo XLSX: {str(e)}")
        return False, errores

# =============================================================================
# ETAPA 2: LECTURA Y DETECCIÓN DE CÓDIGOS
# =============================================================================

def leer_xlsx_normalizado(filepath: Path) -> List[SKUNormalizado]:
    """
    Lee el XLSX y convierte cada fila en un SKUNormalizado
    
    Returns:
        Lista de SKUs normalizados
    """
    skus = []
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb["Normalizado"]
        
        # Leer encabezados
        encabezados = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        
        # Procesar filas de datos
        for fila_idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            
            # Función helper para obtener valor de columna
            def get_col(nombre: str) -> any:
                idx = encabezados.get(nombre)
                if idx and idx <= len(fila):
                    valor = fila[idx - 1]
                    # Convertir None a string vacío para campos de texto
                    if valor is None and nombre not in ['precio_lista', 'precio_neto']:
                        return ''
                    return valor
                return None if nombre in ['precio_lista', 'precio_neto'] else ''
            
            sku = SKUNormalizado(
                codigo_proveedor=str(get_col('codigo_proveedor')),
                descripcion_original=get_col('descripcion_original'),
                fabricante=get_col('fabricante'),
                familia_sugerida=get_col('familia_sugerida'),
                codigo_oem=get_col('codigo_oem'),
                precio_lista=get_col('precio_lista'),
                precio_neto=get_col('precio_neto'),
                aplicaciones=get_col('aplicaciones'),
                tipo_lados=get_col('tipo_lados') or 'sin_lado',
                lado=get_col('lado') or 'N/A',
                carroceria=get_col('carroceria'),
                puertas=get_col('puertas'),
                posicion=get_col('posicion'),
                motor=get_col('motor'),
                caja=get_col('caja'),
                version=get_col('version'),
                observaciones=get_col('observaciones'),
                codigo_padre_proveedor=get_col('codigo_padre_proveedor'),  # Nuevo
                fila_origen=fila_idx
            )
            
            skus.append(sku)
        
        logger.info(f"[OK] Leídos {len(skus)} SKUs del XLSX")
        return skus
        
    except Exception as e:
        logger.error(f"[ERROR] Error leyendo XLSX: {e}")
        raise

# =============================================================================
# ETAPA 3: MATCHEO DE SKU (dupla fabricante + familia)
# =============================================================================

def limpiar_valor_opcional(valor):
    """Convierte strings vacíos y None a NULL para Postgres"""
    if valor is None or valor == "" or valor == "N/A":
        return None
    return valor

def determinar_tipo_fabricante(nombre: str) -> str:
    """
    Determina el tipo de fabricante basándose en el nombre
    
    Returns:
        tipo_fabricante_enum: fabricante_real, etiqueta_origen, etiqueta_categoria, marca_proveedor
    """
    nombre_upper = nombre.upper().strip()
    
    # Etiquetas de origen
    etiquetas_origen = ['IMPORTADO', 'NACIONAL', 'ORIGINAL', 'IMP', 'NAC', 'ORIG']
    if nombre_upper in etiquetas_origen:
        return 'etiqueta_origen'
    
    # Etiquetas de categoría (genéricos contextuales)
    etiquetas_categoria = ['TAIWAN', 'IND.ARG.', 'BRASIL', 'GENÉRICO', 'UNIVERSAL']
    if nombre_upper in etiquetas_categoria:
        return 'etiqueta_categoria'
    
    # Marcas de proveedor conocidas (según reportes COO)
    marcas_proveedor = ['AUTOCLIP', 'SG', 'DELVISO', 'POR IDENTIFICAR']
    if nombre_upper in marcas_proveedor:
        return 'marca_proveedor'
    
    # Por defecto: fabricante real
    return 'fabricante_real'

def generar_codigo_piezauto(db: DatabaseConnection) -> str:
    """
    Genera el próximo código Piezauto secuencial (PZ-000001, PZ-000002, ...)
    
    Returns:
        Código único formato PZ-XXXXXX (6 dígitos)
    """
    query = """
        SELECT codigo_piezauto FROM cat_skus 
        WHERE codigo_piezauto ~ '^PZ-[0-9]+(-[DI])?$'
        ORDER BY 
            CAST(REGEXP_REPLACE(codigo_piezauto, '[^0-9]', '', 'g') AS INTEGER) DESC
        LIMIT 1
    """
    resultado = db.ejecutar_query(query)
    
    if resultado and resultado[0][0]:
        ultimo_codigo = resultado[0][0]
        # Extraer número del formato PZ-000001 o PZ-000001-D
        numero_str = ultimo_codigo.replace('PZ-', '').split('-')[0]
        proximo_numero = int(numero_str) + 1
    else:
        proximo_numero = 1
    
    return f"PZ-{proximo_numero:06d}"

def obtener_o_crear_fabricante(db: DatabaseConnection, nombre: str, tipo: str, familia_id: str) -> str:
    """
    Busca fabricante por (nombre, tipo, familia) o lo crea si no existe
    
    Returns:
        UUID del fabricante
    """
    # Buscar existente
    query = """
        SELECT id FROM cat_fabricantes 
        WHERE nombre = %s AND tipo = %s
        LIMIT 1
    """
    resultado = db.ejecutar_query(query, (nombre, tipo))
    
    if resultado:
        return resultado[0][0]
    
    # Crear nuevo
    query = """
        INSERT INTO cat_fabricantes (nombre, tipo, pais, observaciones)
        VALUES (%s, %s, NULL, %s)
        RETURNING id
    """
    obs = f"Creado automáticamente durante ingesta. Familia: {familia_id}"
    resultado = db.ejecutar_query(query, (nombre, tipo, obs))
    
    fabricante_id = resultado[0][0]
    logger.info(f"[OK] Fabricante creado: {nombre} ({tipo})")
    return fabricante_id

def obtener_o_crear_familia(db: DatabaseConnection, nombre: str) -> str:
    """
    Busca familia por nombre o la crea si no existe
    
    Returns:
        UUID de la familia
    """
    # Buscar existente
    query = "SELECT id FROM cat_familias WHERE nombre = %s"
    resultado = db.ejecutar_query(query, (nombre,))
    
    if resultado:
        return resultado[0][0]
    
    # Crear nueva (con rentabilidad default 25%)
    query = """
        INSERT INTO cat_familias (nombre, rentabilidad_porcentaje)
        VALUES (%s, 25.00)
        RETURNING id
    """
    resultado = db.ejecutar_query(query, (nombre,))
    
    familia_id = resultado[0][0]
    logger.info(f"[OK] Familia creada: {nombre}")
    return familia_id

# =============================================================================
# ETAPA 4: AUTOGENERACIÓN DE LADOS COMBINADOS
# =============================================================================

def autogenerar_lados_combinados(
    db: DatabaseConnection,
    sku_padre: SKUNormalizado,
    fabricante_id: str,
    familia_id: str,
    proveedor_id: str
) -> Tuple[str, str, str]:
    """
    Genera padre + hijo D + hijo I para un SKU con lados_combinados
    
    Returns:
        (padre_id, hijo_d_id, hijo_i_id)
    """
    
    # PASO 1: Crear padre
    query_padre = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, NULL, NULL, TRUE, 'lados_combinados',
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, 'N/A', %s,
            %s, %s, %s, %s,
            TRUE, FALSE
        ) RETURNING id
    """
    
    # Generar código Piezauto único
    codigo_padre = generar_codigo_piezauto(db)
    
    resultado = db.ejecutar_query(query_padre, (
        codigo_padre,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    
    padre_id = resultado[0][0]
    
    # PASO 2: Crear hijo Derecho
    query_hijo = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, %s, %s, FALSE, 'lados_combinados',
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            TRUE, TRUE
        ) RETURNING id
    """
    
    # Hijo D - padre_id (UUID) + codigo_raiz (VARCHAR del código proveedor)
    resultado = db.ejecutar_query(query_hijo, (
        f"{codigo_padre}-D", padre_id, sku_padre.codigo_proveedor,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        'Der', 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    hijo_d_id = resultado[0][0]
    
    # Hijo I - padre_id (UUID) + codigo_raiz (VARCHAR del código proveedor)
    resultado = db.ejecutar_query(query_hijo, (
        f"{codigo_padre}-I", padre_id, sku_padre.codigo_proveedor,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        'Izq', 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    hijo_i_id = resultado[0][0]
    
    # PASO 3 y 4: Crear equivalencias (padre inactivo, hijos activos)
    query_equiv = """
        INSERT INTO cat_equivalencias (
            sku_id, proveedor_id, codigo_proveedor,
            precio_lista_snapshot, precio_neto_snapshot, activo
        ) VALUES (%s, %s, %s, %s, %s, %s)
    """
    
    # Equivalencia padre (inactiva)
    db.ejecutar_query(query_equiv, (
        padre_id, proveedor_id, sku_padre.codigo_proveedor,
        sku_padre.precio_lista, sku_padre.precio_neto, False
    ))
    
    # Equivalencia hijo D (activa) - usar sufijo -D en codigo_proveedor
    db.ejecutar_query(query_equiv, (
        hijo_d_id, proveedor_id, f"{sku_padre.codigo_proveedor}-D",
        sku_padre.precio_lista, sku_padre.precio_neto, True
    ))
    
    # Equivalencia hijo I (activa) - usar sufijo -I en codigo_proveedor
    db.ejecutar_query(query_equiv, (
        hijo_i_id, proveedor_id, f"{sku_padre.codigo_proveedor}-I",
        sku_padre.precio_lista, sku_padre.precio_neto, True
    ))
    
    logger.debug(f"[OK] Lados combinados: {codigo_padre} → {codigo_padre}-D, {codigo_padre}-I")
    
    return padre_id, hijo_d_id, hijo_i_id

def crear_sku_simple(
    db: DatabaseConnection,
    sku: SKUNormalizado,
    fabricante_id: str,
    familia_id: str,
    proveedor_id: str
) -> str:
    """
    Crea un SKU simple (sin desdoblamiento de lados)
    
    Para tipo_lados: juego_indivisible, kit, lado_explicito, sin_lado
    
    Returns:
        sku_id
    """
    
    # Generar código Piezauto único
    codigo_piezauto = generar_codigo_piezauto(db)
    
    # Determinar lado según tipo_lados
    if sku.tipo_lados == 'juego_indivisible':
        lado = 'Ambos'
    elif sku.tipo_lados == 'lado_explicito':
        lado = sku.lado  # Ya viene poblado del XLSX
    else:  # kit, sin_lado
        lado = 'N/A'
    
    # Debug logging detallado
    logger.debug(f"Creando SKU: {sku.codigo_proveedor}")
    logger.debug(f"  tipo_lados: {sku.tipo_lados}, lado: {lado}")
    logger.debug(f"  posicion: '{sku.posicion}' → '{limpiar_valor_opcional(sku.posicion)}'")
    logger.debug(f"  carroceria: '{sku.carroceria}' → '{limpiar_valor_opcional(sku.carroceria)}'")
    logger.debug(f"  puertas: '{sku.puertas}' → '{limpiar_valor_opcional(sku.puertas)}'")
    
    # Insertar SKU
    query = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, NULL, NULL, FALSE, %s,
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            TRUE, TRUE
        ) RETURNING id
    """
    
    resultado = db.ejecutar_query(query, (
        codigo_piezauto, sku.tipo_lados,
        fabricante_id, familia_id, sku.descripcion_original,
        limpiar_valor_opcional(sku.codigo_oem), 
        sku.precio_lista, sku.precio_neto,
        sku.aplicaciones, 
        limpiar_valor_opcional(sku.carroceria), 
        limpiar_valor_opcional(sku.puertas), 
        lado, 
        limpiar_valor_opcional(sku.posicion),
        limpiar_valor_opcional(sku.motor), 
        limpiar_valor_opcional(sku.caja), 
        limpiar_valor_opcional(sku.version), 
        limpiar_valor_opcional(sku.observaciones)
    ))
    
    sku_id = resultado[0][0]
    
    # Crear equivalencia con el código del proveedor
    query_equiv = """
        INSERT INTO cat_equivalencias (
            sku_id, proveedor_id, codigo_proveedor,
            precio_lista_snapshot, precio_neto_snapshot, activo
        ) VALUES (%s, %s, %s, %s, %s, TRUE)
    """
    
    db.ejecutar_query(query_equiv, (
        sku_id, proveedor_id, sku.codigo_proveedor,
        sku.precio_lista, sku.precio_neto
    ))
    
    logger.debug(f"[OK] SKU simple creado: {codigo_piezauto} ({sku.tipo_lados})")
    
    return sku_id

def crear_hijo_pre_armado(
    db: DatabaseConnection,
    sku_hijo: SKUNormalizado,
    fabricante_id: str,
    familia_id: str,
    proveedor_id: str,
    padre_id: str
) -> str:
    """
    Crea un hijo pre-armado de tipo hijo_de_padre_multiple
    
    El hijo ya viene con su propio codigo_proveedor, descripcion, lado, etc.
    Solo necesitamos vincularlo al padre vía padre_id + codigo_raiz
    
    Args:
        padre_id: UUID del padre en cat_skus
        sku_hijo: Datos del hijo desde XLSX
    
    Returns:
        sku_id del hijo
    """
    
    # Generar código Piezauto único
    codigo_piezauto = generar_codigo_piezauto(db)
    
    # Insertar hijo
    query = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, %s, %s, FALSE, 'hijo_de_padre_multiple',
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            TRUE, TRUE
        ) RETURNING id
    """
    
    resultado = db.ejecutar_query(query, (
        codigo_piezauto, padre_id, sku_hijo.codigo_padre_proveedor,  # codigo_raiz = código del padre
        fabricante_id, familia_id, sku_hijo.descripcion_original,
        limpiar_valor_opcional(sku_hijo.codigo_oem), 
        sku_hijo.precio_lista, sku_hijo.precio_neto,
        sku_hijo.aplicaciones, 
        limpiar_valor_opcional(sku_hijo.carroceria), 
        limpiar_valor_opcional(sku_hijo.puertas), 
        sku_hijo.lado, 
        limpiar_valor_opcional(sku_hijo.posicion),
        limpiar_valor_opcional(sku_hijo.motor), 
        limpiar_valor_opcional(sku_hijo.caja), 
        limpiar_valor_opcional(sku_hijo.version), 
        limpiar_valor_opcional(sku_hijo.observaciones)
    ))
    
    sku_id = resultado[0][0]
    
    # Crear equivalencia con el código del proveedor (activa)
    query_equiv = """
        INSERT INTO cat_equivalencias (
            sku_id, proveedor_id, codigo_proveedor,
            precio_lista_snapshot, precio_neto_snapshot, activo
        ) VALUES (%s, %s, %s, %s, %s, TRUE)
    """
    
    db.ejecutar_query(query_equiv, (
        sku_id, proveedor_id, sku_hijo.codigo_proveedor,
        sku_hijo.precio_lista, sku_hijo.precio_neto
    ))
    
    logger.debug(f"[OK] Hijo pre-armado: {codigo_piezauto} (padre: {sku_hijo.codigo_padre_proveedor})")
    
    return sku_id

# =============================================================================
# FUNCIÓN PRINCIPAL DE INGESTA
# =============================================================================

def ingestar_proveedor(filepath: Path, nombre_proveedor: str) -> EstadisticasIngesta:
    """
    Procesa e ingesta un archivo XLSX de un proveedor
    
    Args:
        filepath: Path al archivo XLSX v4
        nombre_proveedor: Nombre del proveedor
    
    Returns:
        Estadísticas de la ingesta
    """
    stats = EstadisticasIngesta()
    db = DatabaseConnection()
    
    try:
        # Conectar a base de datos
        if not db.conectar():
            raise Exception("No se pudo conectar a Supabase")
        
        # ETAPA 1: Validar estructura
        logger.info(f"=== ETAPA 1: Validando estructura de {filepath.name} ===")
        es_valido, errores = validar_estructura_xlsx(filepath)
        if not es_valido:
            stats.errores.extend(errores)
            return stats
        
        # ETAPA 2: Leer SKUs
        logger.info("=== ETAPA 2: Leyendo SKUs del XLSX ===")
        skus = leer_xlsx_normalizado(filepath)
        stats.total_filas = len(skus)
        
        # Obtener o crear proveedor
        query_prov = """
            INSERT INTO cat_proveedores (nombre)
            VALUES (%s)
            ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre
            RETURNING id
        """
        resultado = db.ejecutar_query(query_prov, (nombre_proveedor,))
        proveedor_id = resultado[0][0]
        
        # ETAPA 3-6: Procesar SKUs en dos fases
        logger.info("=== ETAPAS 3-6: Procesando SKUs (dos fases) ===")
        
        # Separar SKUs entre padres/simples y hijos pre-armados
        skus_padres_simples = [s for s in skus if not s.codigo_padre_proveedor]
        skus_hijos_pre_armados = [s for s in skus if s.codigo_padre_proveedor]
        
        logger.info(f"Fase 1: {len(skus_padres_simples)} SKUs padres/simples")
        logger.info(f"Fase 2: {len(skus_hijos_pre_armados)} hijos pre-armados")
        
        # Diccionario para mapear codigo_proveedor → padre_id (UUID)
        mapa_codigo_a_padre_id = {}
        
        # FASE 1: Procesar padres y SKUs simples EN BATCHES
        BATCH_SIZE = 250  # v5: Reducido de 500 a 250 para evitar timeouts Supabase
        total_batches = (len(skus_padres_simples) + BATCH_SIZE - 1) // BATCH_SIZE
        
        logger.info(f"Procesando {len(skus_padres_simples)} SKUs en {total_batches} batches de {BATCH_SIZE}")
        
        for batch_num in range(total_batches):
            inicio = batch_num * BATCH_SIZE
            fin = min((batch_num + 1) * BATCH_SIZE, len(skus_padres_simples))
            batch = skus_padres_simples[inicio:fin]
            
            logger.info(f"[Batch {batch_num + 1}/{total_batches}] Procesando SKUs {inicio + 1} a {fin}")
            
            for sku in batch:
                try:
                    # Obtener o crear familia
                    familia_id = obtener_o_crear_familia(db, sku.familia_sugerida)
                    
                    # Determinar tipo de fabricante (lógica inteligente)
                    tipo_fabricante = determinar_tipo_fabricante(sku.fabricante)
                    
                    # Obtener o crear fabricante
                    fabricante_id = obtener_o_crear_fabricante(
                        db, sku.fabricante, tipo_fabricante, familia_id
                    )
                    
                    # Procesar según tipo_lados
                    if sku.tipo_lados == 'lados_combinados':
                        # Autogenerar padre + hijos D + I
                        padre_id, hijo_d_id, hijo_i_id = autogenerar_lados_combinados(
                            db, sku, fabricante_id, familia_id, proveedor_id
                        )
                        stats.lados_combinados_padres += 1
                        stats.lados_combinados_hijos += 2
                        
                        # Guardar mapeo para fase 2 (por si hay hijos pre-armados de este padre)
                        mapa_codigo_a_padre_id[sku.codigo_proveedor] = padre_id
                        
                    else:
                        # Crear SKU simple (juego_indivisible, kit, lado_explicito, sin_lado)
                        # Puede ser un padre de hijo_de_padre_multiple también
                        sku_id = crear_sku_simple(
                            db, sku, fabricante_id, familia_id, proveedor_id
                        )
                        stats.nuevos_skus += 1
                        
                        # Si es padre de hijo_de_padre_multiple, guardarlo en el mapa
                        if sku.tipo_lados == 'hijo_de_padre_multiple':
                            mapa_codigo_a_padre_id[sku.codigo_proveedor] = sku_id
                    
                    stats.validadas += 1
                    
                except Exception as e:
                    stats.rechazadas += 1
                    stats.errores.append(f"Fila {sku.fila_origen}: {str(e)}")
                    logger.error(f"[ERROR] Error procesando fila {sku.fila_origen}: {e}")
            
            # COMMIT del batch
            try:
                db.commit()
                logger.info(f"[OK] Batch {batch_num + 1}/{total_batches} confirmado ({len(batch)} SKUs)")
                # v5: Delay de 1 segundo entre batches para evitar saturar Supabase
                time.sleep(1)
            except Exception as e:
                db.rollback()
                logger.error(f"[ERROR] Error en commit batch {batch_num + 1}: {e}")
                raise  # Re-raise para que el error se propague
        
        # FASE 2: Procesar hijos pre-armados EN BATCHES
        if skus_hijos_pre_armados:
            total_batches_hijos = (len(skus_hijos_pre_armados) + BATCH_SIZE - 1) // BATCH_SIZE
            logger.info(f"Procesando {len(skus_hijos_pre_armados)} hijos pre-armados en {total_batches_hijos} batches")
            
            for batch_num in range(total_batches_hijos):
                inicio = batch_num * BATCH_SIZE
                fin = min((batch_num + 1) * BATCH_SIZE, len(skus_hijos_pre_armados))
                batch = skus_hijos_pre_armados[inicio:fin]
                
                logger.info(f"[Batch hijos {batch_num + 1}/{total_batches_hijos}] Procesando {inicio + 1} a {fin}")
                
                for sku_hijo in batch:
                    try:
                        # Buscar el padre en el mapa
                        if sku_hijo.codigo_padre_proveedor not in mapa_codigo_a_padre_id:
                            # Padre no encontrado en esta carga - buscar en base de datos
                            query_buscar_padre = """
                                SELECT s.id 
                                FROM cat_skus s
                                JOIN cat_equivalencias e ON e.sku_id = s.id
                                WHERE e.codigo_proveedor = %s 
                                  AND e.proveedor_id = %s
                                  AND s.es_padre = TRUE
                                LIMIT 1
                            """
                            resultado = db.ejecutar_query(
                                query_buscar_padre, 
                                (sku_hijo.codigo_padre_proveedor, proveedor_id)
                            )
                            
                            if not resultado:
                                raise Exception(f"Padre no encontrado: {sku_hijo.codigo_padre_proveedor}")
                            
                            padre_id = resultado[0][0]
                        else:
                            padre_id = mapa_codigo_a_padre_id[sku_hijo.codigo_padre_proveedor]
                        
                        # Obtener o crear familia y fabricante (puede ser distinto al del padre)
                        familia_id = obtener_o_crear_familia(db, sku_hijo.familia_sugerida)
                        tipo_fabricante = determinar_tipo_fabricante(sku_hijo.fabricante)
                        fabricante_id = obtener_o_crear_fabricante(
                            db, sku_hijo.fabricante, tipo_fabricante, familia_id
                        )
                        
                        # Crear hijo pre-armado
                        hijo_id = crear_hijo_pre_armado(
                            db, sku_hijo, fabricante_id, familia_id, proveedor_id, padre_id
                        )
                        
                        stats.nuevos_skus += 1
                        stats.validadas += 1
                        
                    except Exception as e:
                        stats.rechazadas += 1
                        stats.errores.append(f"Fila {sku_hijo.fila_origen} (hijo): {str(e)}")
                        logger.error(f"[ERROR] Error procesando hijo fila {sku_hijo.fila_origen}: {e}")
                
                # COMMIT del batch de hijos
                try:
                    db.commit()
                    logger.info(f"[OK] Batch hijos {batch_num + 1}/{total_batches_hijos} confirmado")
                    # v5: Delay de 1 segundo entre batches
                    time.sleep(1)
                except Exception as e:
                    db.rollback()
                    logger.error(f"[ERROR] Error en commit batch hijos {batch_num + 1}: {e}")
                    raise
        
        logger.info("[OK] Ingesta completada (commits por batch)")
        
    except Exception as e:
        db.rollback()
        logger.error(f"[ERROR] Error en ingesta: {e}")
        stats.errores.append(str(e))
    finally:
        db.cerrar()
    
    return stats

# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Importador de catálogo Piezauto')
    parser.add_argument('--proveedor', required=True, help='Nombre del proveedor')
    parser.add_argument('--archivo', required=True, help='Path al archivo XLSX v4')
    
    args = parser.parse_args()
    
    filepath = Path(args.archivo)
    if not filepath.exists():
        logger.error(f"[ERROR] Archivo no encontrado: {filepath}")
        sys.exit(1)
    
    logger.info(f"=== INICIANDO INGESTA: {args.proveedor} ===")
    logger.info(f"Archivo: {filepath}")
    
    stats = ingestar_proveedor(filepath, args.proveedor)
    
    # Reporte final
    logger.info("=== RESUMEN DE INGESTA ===")
    logger.info(f"Total filas procesadas: {stats.total_filas}")
    logger.info(f"Validadas: {stats.validadas}")
    logger.info(f"Rechazadas: {stats.rechazadas}")
    logger.info(f"Lados combinados (padres): {stats.lados_combinados_padres}")
    logger.info(f"Lados combinados (hijos): {stats.lados_combinados_hijos}")
    
    if stats.errores:
        logger.warning(f"Errores encontrados: {len(stats.errores)}")
        for error in stats.errores[:10]:  # Mostrar primeros 10
            logger.warning(f"  - {error}")

if __name__ == "__main__":
    main()
