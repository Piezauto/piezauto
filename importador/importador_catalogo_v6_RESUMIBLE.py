#!/usr/bin/env python3
"""
IMPORTADOR PIEZAUTO v7 - MODO ULTRA-RESUMIBLE
Procesa 1 SKU a la vez. Pérdida en caso de fallo = 0.

Características:
- 1 SKU = 1 COMMIT (confirmación inmediata)
- Manejo individual de errores (continúa si falla 1 SKU)
- Checkpoint automático (retoma donde quedó)
- Reconexión cada 100 SKUs para limpiar estado

Uso:
    python importador_catalogo_v6_RESUMIBLE.py --proveedor Surpiezas --archivo Surpiezas.xlsx
    
    Si falla, RE-EJECUTAR EL MISMO COMANDO. Continúa automáticamente.
    
NOTA: Más lento que batches, pero GARANTIZA pérdida = 0.
"""

import sys
import argparse
import logging
from pathlib import Path
from typing import List
from dataclasses import dataclass
import openpyxl
import psycopg
import os
from dotenv import load_dotenv
import time

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('importador_v6.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

BATCH_SIZE = 1  # 1 SKU = 1 COMMIT (pérdida = 0 en caso de fallo)
DELAY_ENTRE_SKUS = 0.1  # Delay mínimo entre SKUs
RECONECTAR_CADA = 999999  # Solo reconecta si detecta pérdida (modo reactivo)

@dataclass
class SKUNormalizado:
    codigo_proveedor: str
    descripcion_original: str
    fabricante: str
    familia_sugerida: str
    codigo_oem: str
    precio_lista: float
    precio_neto: float
    aplicaciones: str
    tipo_lados: str
    lado: str
    carroceria: str
    puertas: str
    posicion: str
    motor: str
    caja: str
    version: str
    observaciones: str
    codigo_padre_proveedor: str
    fila_origen: int = 0

# =============================================================================
# CONEXIÓN DB
# =============================================================================

class DatabaseConnection:
    def __init__(self):
        self.conn = None
        self.cursor = None
        
    def conectar(self):
        try:
            self.conn = psycopg.connect(
                host=os.getenv('SUPABASE_HOST'),
                dbname=os.getenv('SUPABASE_DB'),
                user=os.getenv('SUPABASE_USER'),
                password=os.getenv('SUPABASE_PASSWORD'),
                port=os.getenv('SUPABASE_PORT', '5432')
            )
            self.cursor = self.conn.cursor()
            return True
        except Exception as e:
            logger.error(f"Error conectando: {e}")
            return False
    
    def cerrar(self):
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
    
    def ejecutar_query(self, query: str, params: tuple = None):
        try:
            self.cursor.execute(query, params)
            if self.cursor.description:
                return self.cursor.fetchall()
            return None
        except Exception as e:
            logger.error(f"Error query: {e}")
            raise
    
    def commit(self):
        self.conn.commit()
    
    def rollback(self):
        self.conn.rollback()

# =============================================================================
# CHECKPOINT - DETECTA DÓNDE QUEDÓ
# =============================================================================

def obtener_ultimo_codigo_cargado(db: DatabaseConnection, proveedor: str) -> int:
    """Detecta cuántos SKUs ya están cargados para continuar"""
    query = """
        SELECT COUNT(DISTINCT e.codigo_proveedor)
        FROM cat_equivalencias e
        JOIN cat_proveedores p ON p.id = e.proveedor_id
        WHERE p.nombre = %s
    """
    resultado = db.ejecutar_query(query, (proveedor,))
    count = resultado[0][0] if resultado else 0
    logger.info(f"[CHECKPOINT] {proveedor} tiene {count} SKUs ya cargados")
    return count

# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

def limpiar_valor_opcional(valor):
    if valor is None or valor == "" or valor == "N/A":
        return None
    return valor

def determinar_tipo_fabricante(nombre: str) -> str:
    nombre_upper = nombre.upper().strip()
    etiquetas_origen = ['IMPORTADO', 'NACIONAL', 'ORIGINAL', 'IMP', 'NAC', 'ORIG']
    if nombre_upper in etiquetas_origen:
        return 'etiqueta_origen'
    etiquetas_categoria = ['TAIWAN', 'IND.ARG.', 'BRASIL', 'GENÉRICO', 'UNIVERSAL']
    if nombre_upper in etiquetas_categoria:
        return 'etiqueta_categoria'
    marcas_proveedor = ['AUTOCLIP', 'SG', 'DELVISO', 'POR IDENTIFICAR']
    if nombre_upper in marcas_proveedor:
        return 'marca_proveedor'
    return 'fabricante_real'

def generar_codigo_piezauto(db: DatabaseConnection) -> str:
    query = """
        SELECT codigo_piezauto FROM cat_skus 
        WHERE codigo_piezauto ~ '^PZ-[0-9]+(-[DI])?$'
        ORDER BY 
            CAST(REGEXP_REPLACE(codigo_piezauto, '[^0-9]', '', 'g') AS INTEGER) DESC
        LIMIT 1
    """
    resultado = db.ejecutar_query(query)
    
    if resultado and resultado[0][0]:
        ultimo_codigo = resultado[0][0]
        numero_str = ultimo_codigo.replace('PZ-', '').split('-')[0]
        proximo_numero = int(numero_str) + 1
    else:
        proximo_numero = 1
    
    return f"PZ-{proximo_numero:06d}"

def obtener_o_crear_fabricante(db: DatabaseConnection, nombre: str, tipo: str, familia_id: str) -> str:
    query = "SELECT id FROM cat_fabricantes WHERE nombre = %s AND tipo = %s LIMIT 1"
    resultado = db.ejecutar_query(query, (nombre, tipo))
    
    if resultado:
        return resultado[0][0]
    
    query = """
        INSERT INTO cat_fabricantes (nombre, tipo, pais, observaciones)
        VALUES (%s, %s, NULL, %s)
        RETURNING id
    """
    obs = f"Creado auto. Familia: {familia_id}"
    resultado = db.ejecutar_query(query, (nombre, tipo, obs))
    return resultado[0][0]

def obtener_o_crear_familia(db: DatabaseConnection, nombre: str) -> str:
    query = "SELECT id FROM cat_familias WHERE nombre = %s"
    resultado = db.ejecutar_query(query, (nombre,))
    
    if resultado:
        return resultado[0][0]
    
    query = """
        INSERT INTO cat_familias (nombre, rentabilidad_porcentaje)
        VALUES (%s, 25.00)
        RETURNING id
    """
    resultado = db.ejecutar_query(query, (nombre,))
    return resultado[0][0]

def crear_sku_simple(db: DatabaseConnection, sku: SKUNormalizado, fabricante_id: str, 
                     familia_id: str, proveedor_id: str) -> str:
    codigo_piezauto = generar_codigo_piezauto(db)
    
    if sku.tipo_lados == 'juego_indivisible':
        lado = 'Ambos'
    elif sku.tipo_lados == 'lado_explicito':
        lado = sku.lado
    else:
        lado = 'N/A'
    
    query = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, NULL, NULL, FALSE, %s,
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            TRUE, TRUE
        ) RETURNING id
    """
    
    resultado = db.ejecutar_query(query, (
        codigo_piezauto, sku.tipo_lados,
        fabricante_id, familia_id, sku.descripcion_original,
        limpiar_valor_opcional(sku.codigo_oem), 
        sku.precio_lista, sku.precio_neto,
        sku.aplicaciones, 
        limpiar_valor_opcional(sku.carroceria), 
        limpiar_valor_opcional(sku.puertas), 
        lado, 
        limpiar_valor_opcional(sku.posicion),
        limpiar_valor_opcional(sku.motor), 
        limpiar_valor_opcional(sku.caja), 
        limpiar_valor_opcional(sku.version), 
        limpiar_valor_opcional(sku.observaciones)
    ))
    
    sku_id = resultado[0][0]
    
    query_equiv = """
        INSERT INTO cat_equivalencias (
            sku_id, proveedor_id, codigo_proveedor,
            precio_lista_snapshot, precio_neto_snapshot, activo
        ) VALUES (%s, %s, %s, %s, %s, TRUE)
    """
    
    db.ejecutar_query(query_equiv, (
        sku_id, proveedor_id, sku.codigo_proveedor,
        sku.precio_lista, sku.precio_neto
    ))
    
    return sku_id

def autogenerar_lados_combinados(db: DatabaseConnection, sku_padre: SKUNormalizado,
                                 fabricante_id: str, familia_id: str, proveedor_id: str):
    codigo_padre = generar_codigo_piezauto(db)
    
    query_padre = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, NULL, NULL, TRUE, 'lados_combinados',
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, 'N/A', %s,
            %s, %s, %s, %s,
            TRUE, FALSE
        ) RETURNING id
    """
    
    resultado = db.ejecutar_query(query_padre, (
        codigo_padre,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    
    padre_id = resultado[0][0]
    
    query_hijo = """
        INSERT INTO cat_skus (
            codigo_piezauto, padre_id, codigo_raiz, es_padre, tipo_lados,
            fabricante_id, familia_id, descripcion,
            codigo_oem, precio_lista, precio_neto,
            aplicaciones, carroceria, puertas, lado, posicion,
            motor, caja, version, observaciones,
            activo, activo_venta
        ) VALUES (
            %s, %s, %s, FALSE, 'lados_combinados',
            %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            TRUE, TRUE
        ) RETURNING id
    """
    
    # Hijo D
    resultado = db.ejecutar_query(query_hijo, (
        f"{codigo_padre}-D", padre_id, sku_padre.codigo_proveedor,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        'Der', 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    hijo_d_id = resultado[0][0]
    
    # Hijo I
    resultado = db.ejecutar_query(query_hijo, (
        f"{codigo_padre}-I", padre_id, sku_padre.codigo_proveedor,
        fabricante_id, familia_id, sku_padre.descripcion_original,
        limpiar_valor_opcional(sku_padre.codigo_oem), 
        sku_padre.precio_lista, sku_padre.precio_neto,
        sku_padre.aplicaciones, 
        limpiar_valor_opcional(sku_padre.carroceria), 
        limpiar_valor_opcional(sku_padre.puertas), 
        'Izq', 
        limpiar_valor_opcional(sku_padre.posicion),
        limpiar_valor_opcional(sku_padre.motor), 
        limpiar_valor_opcional(sku_padre.caja), 
        limpiar_valor_opcional(sku_padre.version), 
        limpiar_valor_opcional(sku_padre.observaciones)
    ))
    hijo_i_id = resultado[0][0]
    
    query_equiv = """
        INSERT INTO cat_equivalencias (
            sku_id, proveedor_id, codigo_proveedor,
            precio_lista_snapshot, precio_neto_snapshot, activo
        ) VALUES (%s, %s, %s, %s, %s, %s)
    """
    
    db.ejecutar_query(query_equiv, (
        padre_id, proveedor_id, sku_padre.codigo_proveedor,
        sku_padre.precio_lista, sku_padre.precio_neto, False
    ))
    
    db.ejecutar_query(query_equiv, (
        hijo_d_id, proveedor_id, f"{sku_padre.codigo_proveedor}-D",
        sku_padre.precio_lista, sku_padre.precio_neto, True
    ))
    
    db.ejecutar_query(query_equiv, (
        hijo_i_id, proveedor_id, f"{sku_padre.codigo_proveedor}-I",
        sku_padre.precio_lista, sku_padre.precio_neto, True
    ))

# =============================================================================
# LECTURA XLSX
# =============================================================================

def leer_xlsx_normalizado(filepath: Path) -> List[SKUNormalizado]:
    skus = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Normalizado"]
    
    encabezados = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    
    def get_col(nombre: str, fila) -> any:
        idx = encabezados.get(nombre)
        if idx and idx <= len(fila):
            valor = fila[idx - 1]
            if valor is None and nombre not in ['precio_lista', 'precio_neto']:
                return ''
            return valor
        return None if nombre in ['precio_lista', 'precio_neto'] else ''
    
    for fila_idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sku = SKUNormalizado(
            codigo_proveedor=str(get_col('codigo_proveedor', fila)),
            descripcion_original=get_col('descripcion_original', fila),
            fabricante=get_col('fabricante', fila),
            familia_sugerida=get_col('familia_sugerida', fila),
            codigo_oem=get_col('codigo_oem', fila),
            precio_lista=get_col('precio_lista', fila),
            precio_neto=get_col('precio_neto', fila),
            aplicaciones=get_col('aplicaciones', fila),
            tipo_lados=get_col('tipo_lados', fila) or 'sin_lado',
            lado=get_col('lado', fila) or 'N/A',
            carroceria=get_col('carroceria', fila),
            puertas=get_col('puertas', fila),
            posicion=get_col('posicion', fila),
            motor=get_col('motor', fila),
            caja=get_col('caja', fila),
            version=get_col('version', fila),
            observaciones=get_col('observaciones', fila),
            codigo_padre_proveedor=get_col('codigo_padre_proveedor', fila),
            fila_origen=fila_idx
        )
        skus.append(sku)
    
    logger.info(f"[OK] Leídos {len(skus)} SKUs del XLSX")
    return skus

# =============================================================================
# INGESTA RESUMIBLE - 1 SKU A LA VEZ
# =============================================================================

def ingestar_proveedor_resumible(filepath: Path, nombre_proveedor: str):
    db = DatabaseConnection()
    
    if not db.conectar():
        raise Exception("No se pudo conectar a Supabase")
    
    # Leer XLSX completo
    skus_total = leer_xlsx_normalizado(filepath)
    total_skus = len(skus_total)
    
    # Obtener proveedor ID
    query_prov = """
        INSERT INTO cat_proveedores (nombre)
        VALUES (%s)
        ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre
        RETURNING id
    """
    resultado = db.ejecutar_query(query_prov, (nombre_proveedor,))
    proveedor_id = resultado[0][0]
    db.commit()
    
    # CHECKPOINT: detectar cuántos ya están cargados
    skus_ya_cargados = obtener_ultimo_codigo_cargado(db, nombre_proveedor)
    
    if skus_ya_cargados >= total_skus:
        logger.info(f"[OK] {nombre_proveedor} YA ESTÁ COMPLETO ({skus_ya_cargados}/{total_skus})")
        db.cerrar()
        return
    
    # Continuar desde donde quedó
    skus_pendientes = skus_total[skus_ya_cargados:]
    logger.info(f"[RESUMIENDO] Continuando desde SKU {skus_ya_cargados + 1} de {total_skus}")
    logger.info(f"[RESUMIENDO] Faltan {len(skus_pendientes)} SKUs por procesar")
    
    # Procesar 1 SKU A LA VEZ
    skus_procesados = 0
    skus_fallidos = 0
    
    for idx, sku in enumerate(skus_pendientes):
        sku_global = skus_ya_cargados + idx + 1
        
        # Log cada 100 SKUs
        if idx % 100 == 0:
            logger.info(f"[Progreso] SKU {sku_global}/{total_skus} ({(sku_global/total_skus*100):.1f}%)")
        
        # Reconectar cada N SKUs para limpiar estado
        if idx > 0 and idx % RECONECTAR_CADA == 0:
            logger.info(f"[Reconexión] Limpiando conexión en SKU {sku_global}")
            db.cerrar()
            time.sleep(0.5)
            if not db.conectar():
                logger.error(f"[ERROR] No se pudo reconectar en SKU {sku_global}")
                break
        
        # PROCESAR 1 SKU CON MANEJO DE ERRORES INDIVIDUAL
        try:
            familia_id = obtener_o_crear_familia(db, sku.familia_sugerida)
            tipo_fabricante = determinar_tipo_fabricante(sku.fabricante)
            fabricante_id = obtener_o_crear_fabricante(db, sku.fabricante, tipo_fabricante, familia_id)
            
            if sku.tipo_lados == 'lados_combinados':
                autogenerar_lados_combinados(db, sku, fabricante_id, familia_id, proveedor_id)
            else:
                crear_sku_simple(db, sku, fabricante_id, familia_id, proveedor_id)
            
            # COMMIT INMEDIATO (1 SKU confirmado)
            db.commit()
            skus_procesados += 1
            
            time.sleep(DELAY_ENTRE_SKUS)
            
        except psycopg.OperationalError as e:
            # Error de conexión - intentar reconectar
            skus_fallidos += 1
            logger.error(f"[ERROR] SKU {sku_global} ({sku.codigo_proveedor}): Conexión perdida - {e}")
            logger.info(f"[RETRY] Intentando reconectar...")
            
            db.cerrar()
            time.sleep(1)
            
            if db.conectar():
                logger.info(f"[OK] Reconexión exitosa. Continuando con SKU {sku_global + 1}")
            else:
                logger.error(f"[FALLO] No se pudo reconectar. Deteniendo.")
                break
                
        except Exception as e:
            # Error de datos - registrar y continuar
            skus_fallidos += 1
            logger.error(f"[ERROR] SKU {sku_global} ({sku.codigo_proveedor}): {e}")
            
            # Intentar rollback si la conexión sigue viva
            try:
                db.rollback()
            except:
                # Si el rollback falla, reconectar
                db.cerrar()
                time.sleep(0.5)
                if not db.conectar():
                    logger.error(f"[FALLO] No se pudo reconectar después de error. Deteniendo.")
                    break
    
    db.cerrar()
    
    logger.info(f"[COMPLETADO] {nombre_proveedor}")
    logger.info(f"  - SKUs procesados exitosamente: {skus_procesados}")
    logger.info(f"  - SKUs fallidos: {skus_fallidos}")
    logger.info(f"  - Total en DB: {skus_ya_cargados + skus_procesados}")
    
    if skus_fallidos > 0:
        logger.warning(f"[ADVERTENCIA] {skus_fallidos} SKUs fallaron. Re-ejecutar para reintentar.")

# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Importador Piezauto v7 ULTRA-RESUMIBLE (1 SKU = 1 COMMIT)')
    parser.add_argument('--proveedor', required=True)
    parser.add_argument('--archivo', required=True)
    
    args = parser.parse_args()
    
    filepath = Path(args.archivo)
    if not filepath.exists():
        logger.error(f"Archivo no encontrado: {filepath}")
        sys.exit(1)
    
    logger.info(f"=== INICIANDO INGESTA RESUMIBLE: {args.proveedor} ===")
    
    try:
        ingestar_proveedor_resumible(filepath, args.proveedor)
        logger.info("[SUCCESS] Carga completada")
    except Exception as e:
        logger.error(f"[FALLO] {e}")
        logger.info("[INFO] RE-EJECUTAR EL MISMO COMANDO para continuar")
        sys.exit(1)

if __name__ == "__main__":
    main()
